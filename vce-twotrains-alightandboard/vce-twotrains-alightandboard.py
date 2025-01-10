"""
This is a recursive peak-hour platform clearance calculator.
model from https://onlinepubs.trb.org/Onlinepubs/hrr/1971/355/355-001.pdf
"""

import dataclasses
import numpy as np
import openpyxl
from openpyxl.cell import Cell
from openpyxl.chart import Reference, Series

# basic flow: train egress > platform crowd > VCE egress rate > back to
# platform crowd


# keep high VCE egress rate if queues at stairs are long
def alight_rate_fn(k, t, t0, u):
    """
    :param k: number of people waiting to get off train
    :param t: time pass counter (s)
    :param t0: train arrival time
    :param u: train(x)doors*rate (1 pax/door/s)
    :return: egress rate from train to platform across all doors (pax/s)
    """
    if t > t0:
        return min(k, u)
    else:
        return 0


def plat_clearance_fn(karr, a, w, qmax):
    """
    :param a: usable platform area
    :param w: total width of vertical circulation elements
    :param: karr: number of people waiting to get onto a stairwell
    :param: qmax: number of people that can fit around stair thresholds
    :return: platform egress rate on stairs
    """
    if karr <= qmax:
        return min(
            karr,
            min(17 * w / 60, (111 * a / max(1, karr) - 162) / (a / max(1, karr)) ** 2),
        )
    else:
        return max(
            10 * w / 60,
            # Min of upstairs LOS C/D boundary flow rate
            min(17 * w / 60, (111 * a / max(1, karr) - 162) / (a / max(1, karr)) ** 2),
        )


def plat_ingress_fn(kdep, a, w, r_up):
    """
    :param kdep: number of people waiting to get onto a stairwell
    :param a: usable concourse area
    :param w: total width of vertical circulation elements
    :param: r_up: upstairs flow, passed from plat_egress_fn
    :return: platform ingress rate on stairs
    """
    # 1st question, how much downstairs flow demand exists?
    # 2nd question, how much stair capacity does upstairs flow take?
    # P = (111M - 162)/(M^2) is the upstairs flow eq per ft wide.
    if kdep > 0:
        return min(
            kdep,
            min(
                max(0, 12 * w / 60 - r_up),
                # Max of downstairs LOS C/D boundary flow rate
                max(
                    0,
                    (
                        (111 * (a / max(1, kdep)) - 162) / ((a / max(1, kdep)) ** 2)
                        - r_up
                    ),
                ),
            ),
        )
    else:
        return 0


def boarder_frac_fn(trainA_boarders, trainB_boarders):
    if trainA_boarders + trainB_boarders > 0:
        return trainA_boarders / (trainB_boarders + trainA_boarders)
    else:
        return 1


def board_rate_fn(r_max, r_off, sim_t, arr_t, dep_t, boarders):
    """
    :param vmax: maximum train deboard rate,
    pass train1_doors or train2_doors from params, since 1 door/sec
    :param r_off: train alight rate, pass alight_rate_fn
    :param sim_t: time in seconds, pass counter
    :param arr_t: train arrival time
    :param: dep_t: train departure time
    :param: boarders: number of passengers waiting on platform to board,
    pass departing_pax_on_plat
    :return: train ingress rate across all doors (pax/s)
    """
    if arr_t < sim_t < dep_t:
        return min(r_max - r_off, boarders)
    else:
        return 0


def space_per_pax_fn(k, a):
    """
    :param k: people on platform (pax)
    :param a: usable platform area (ft^2)
    :return: space per passenger (ft^2/pax)
    """
    if k > 0:
        return a / k
    else:
        return a


def plat_crowd_grade(inst_crowding: float) -> str:
    if inst_crowding > 35:
        return "A"
    elif 25 < inst_crowding <= 35:
        return "B"
    elif 15 < inst_crowding <= 25:
        return "C"
    elif 10 < inst_crowding <= 15:
        return "D"
    elif 5 < inst_crowding <= 10:
        return "E"
    else:
        return "F"


def egress_crowd_grade(w, plat_egress_rate: float) -> str:
    if plat_egress_rate <= w * 5 / 60:
        return "A"
    elif w * 5 / 60 < plat_egress_rate <= w * 7 / 60:
        return "B"
    elif w * 7 / 60 < plat_egress_rate <= w * 9.5 / 60:
        return "C"
    elif w * 9.5 / 60 < plat_egress_rate <= w * 13 / 60:
        return "D"
    elif w * 13 / 60 < plat_egress_rate <= w * 17 / 60:
        return "E"
    else:
        return "F"


@dataclasses.dataclass
class Params:
    # File slug
    file_slug: str

    # how many seconds we want to simulate
    simulation_time: int

    width: int

    length: int

    # area correction factor
    cf: float

    # train 1 arriving passenger load
    train1_arriving_pax: int

    # train 2 arriving passenger load
    train2_arriving_pax: int

    # train 1 departing passenger demand
    train1_outbound_demand: int

    # train 2 departing passenger demand
    train2_outbound_demand: int

    # single-door equivalents
    train1_doors: int

    # single-door equivalents
    train2_doors: int

    # time of first train arrival
    arr_time1: int

    # time of second train arrival
    arr_time2: int

    # feet of queue in front of each stair that pushes max flow
    queue_length: int

    # total width of upstairs VCEs in feet.
    w: float

    ww: any

    # Number of people already on platform at time 0 wanting to board train 1
    train1_awaiting_boarders: int

    # Number of people already on platform at time 0 wanting to board train 2
    train2_awaiting_boarders: int


def calc_workbook(params: Params) -> openpyxl.Workbook:
    eff_area = params.width * params.length * params.cf

    www = params.ww[0, :]

    print("www = ", www)

    # Initialize counters
    total_pax_on_platform = 0
    arrived_pax_waiting_on_plat = 0
    train1_remaining_arrivals = params.train1_arriving_pax
    train2_remaining_arrivals = params.train2_arriving_pax
    train1_new_pax = 0
    train2_new_pax = 0
    train1_boarders_upstairs = (
        params.train1_outbound_demand - params.train1_awaiting_boarders
    )
    train2_boarders_upstairs = (
        params.train2_outbound_demand - params.train2_awaiting_boarders
    )
    train1_boarders_on_plat = params.train1_awaiting_boarders
    train2_boarders_on_plat = params.train2_awaiting_boarders
    total_pax_on_platform = (
        params.train1_awaiting_boarders + params.train2_awaiting_boarders
    )
    wb = openpyxl.Workbook()

    sheet = wb.active

    rownum = 0

    def make_row(value, description: str):
        nonlocal rownum
        rownum = rownum + 1
        sheet.cell(column=1, row=rownum).value = description
        sheet.cell(column=2, row=rownum).value = value

    make_row("Value", "Parameter")
    make_row(params.width, "Platform width (ft)")
    make_row(params.length, "Platform length (ft)")
    make_row(params.w, "Total VCE width (ft)")
    make_row(params.cf, "Effective Area Multiplier")
    make_row(eff_area, "Usable Platform Area (sqft)")
    make_row(params.train1_arriving_pax, "Train 1 Arriving Passengers")
    make_row(params.train1_outbound_demand, "Train 1 Departing Passengers")
    make_row(params.arr_time1, "Train 1 Arrival Time")
    make_row(params.train2_arriving_pax, "Train 2 Arriving Passengers")
    make_row(params.train2_outbound_demand, "Train 2 Departing Passengers")
    make_row(params.arr_time2, "Train 2 Arrival Time")
    make_row(params.simulation_time, "Simulation Length (s)")
    make_row(params.w * 19 / 60, "LOS F Egress Rate (pax/s)")
    make_row(
        (params.train1_arriving_pax + params.train2_arriving_pax)
        / (params.w * 19 / 60),
        "Emergency Egress Time (s)",
    )

    del rownum

    class Columns:
        pass

    columns = Columns()
    # The input parameters go in a table taking up columns 0 (A) and 1 (B).
    colnum = 2

    def make_column_num(description: str):
        nonlocal colnum
        colnum = colnum + 1
        sheet.cell(row=1, column=colnum).value = description
        return colnum

    columns.time_after = make_column_num("Time after arrival (s)")
    columns.train1_pax = make_column_num("Passengers on Train 1")
    columns.train2_pax = make_column_num("Passengers on Train 2")
    columns.train1_off_rate = make_column_num("Train 1 Alight Rate (pax/s)")
    columns.train2_off_rate = make_column_num("Train 2 Alight Rate (pax/s)")
    columns.train1_on_rate = make_column_num("Train 1 Board Rate (pax/s)")
    columns.train2_on_rate = make_column_num("Train 2 Board Rate (pax/s)")
    columns.down_rate = make_column_num("Downstairs Rate (pax/s)")
    columns.up_rate = make_column_num("Upstairs Rate (pax/s)")
    columns.departing_pax_on_plat_1 = make_column_num(
        "Train 1 Departing Passengers on Platform"
    )
    columns.departing_pax_on_plat_2 = make_column_num(
        "Train 2 Departing Passengers on Platform"
    )
    columns.arrived_pax_waiting_on_plat = make_column_num(
        "Arrived Passengers on Platform"
    )
    columns.total_pax_on_platform = make_column_num("Total Passengers on Platform")
    columns.inst_crowding = make_column_num("Platform Space per Passanger (sqft)")
    columns.net_pax_flow_rate = make_column_num("Net Platform Flow Rate")
    columns.plat_crowd_los = make_column_num("Platform Crowding LOS")
    columns.egress_los = make_column_num("Egress LOS")

    del colnum

    FIRST_DATA_ROW = 2

    print("Elapsed_Time", "Train_1_Pax", "Train_2_Pax")

    for time_after in range(0, params.simulation_time):
        train1_off_rate = alight_rate_fn(
            train1_remaining_arrivals,
            time_after,
            params.arr_time1,
            params.train1_doors,
        )
        train1_remaining_arrivals -= train1_off_rate
        if train1_remaining_arrivals < 0:
            train1_remaining_arrivals = 0
        train2_off_rate = alight_rate_fn(
            train2_remaining_arrivals,
            time_after,
            params.arr_time2,
            params.train2_doors,
        )
        train2_remaining_arrivals -= train2_off_rate
        if train2_remaining_arrivals < 0:
            train2_remaining_arrivals = 0
        total_pax_on_platform = (
            total_pax_on_platform + train1_off_rate + train2_off_rate
        )
        arrived_pax_waiting_on_plat = (
            arrived_pax_waiting_on_plat + train1_off_rate + train2_off_rate
        )
        plat_egress_rate = plat_clearance_fn(
            arrived_pax_waiting_on_plat,
            eff_area,
            params.w,
            params.w * params.queue_length / 5,
        )
        arrived_pax_waiting_on_plat -= plat_egress_rate
        if arrived_pax_waiting_on_plat < 0:
            arrived_pax_waiting_on_plat = 0
        total_pax_on_platform -= plat_egress_rate
        plat_ingress_rate_1 = plat_ingress_fn(
            train1_boarders_upstairs,
            5000,
            params.w
            * boarder_frac_fn(train1_boarders_upstairs, train2_boarders_upstairs),
            plat_egress_rate,
        )

        plat_ingress_rate_2 = plat_ingress_fn(
            train2_boarders_upstairs,
            5000,
            params.w
            * boarder_frac_fn(train2_boarders_upstairs, train1_boarders_upstairs),
            plat_egress_rate,
        )
        train1_boarders_on_plat += plat_ingress_rate_1
        train2_boarders_on_plat += plat_ingress_rate_2
        total_pax_on_platform += plat_ingress_rate_1
        total_pax_on_platform += plat_ingress_rate_2
        train1_on_rate = board_rate_fn(
            params.train1_doors,
            train1_off_rate,
            time_after,
            params.arr_time1,
            params.simulation_time,
            train1_boarders_on_plat,
        )
        train2_on_rate = board_rate_fn(
            params.train2_doors,
            train2_off_rate,
            time_after,
            params.arr_time2,
            params.simulation_time,
            train2_boarders_on_plat,
        )

        train1_boarders_on_plat -= train1_on_rate

        train2_boarders_on_plat -= train2_on_rate

        total_pax_on_platform -= train1_on_rate

        total_pax_on_platform -= train2_on_rate

        train1_boarders_upstairs -= plat_ingress_rate_1

        train2_boarders_upstairs -= plat_ingress_rate_2

        train1_new_pax += train1_on_rate

        train2_new_pax += train2_on_rate

        inst_crowding = space_per_pax_fn(total_pax_on_platform, eff_area)
        if total_pax_on_platform < 0:
            total_pax_on_platform = 0
        if train1_boarders_on_plat < 0:
            train1_boarders_on_plat = 0
        if train2_boarders_on_plat < 0:
            train2_boarders_on_plat = 0
        if arrived_pax_waiting_on_plat < 0:
            arrived_pax_waiting_on_plat = 0
        print(
            time_after,
            train1_remaining_arrivals + train1_new_pax,
            train2_remaining_arrivals + train2_new_pax,
            arrived_pax_waiting_on_plat,
            plat_egress_rate,
        )
        """
        print(
            "At time " + str(time_after) + " s,",
            str(train1_remaining_arrivals) + " wait to alight train 1;",
            str(train2_remaining_arrivals) + " wait to alight train 2;",
            str(train1_on_rate - train1_off_rate) + " pax/s train 1 net rate;",
            str(train2_on_rate - train2_off_rate) + " pax/s train 2 net rate;",
        )
        print(
            str(int(arrived_pax_waiting_on_plat))
            + " deboarded pax on platform;",
            str(int(train1_boarders_on_plat + int(train2_boarders_on_plat)))
            + " boarding pax on platform;",
            str(int(total_pax_on_platform)) + " total pax on platform;",
            str(int(inst_crowding)) + " sqft per pax;",
        )
        print(
            str(plat_egress_rate) + " pax/s up;",
            str((plat_ingress_rate_1 + plat_ingress_rate_2)) + "pax/s down;",
            str((train1_boarders_upstairs + train2_boarders_upstairs))
            + " pax are upstairs"
        )
        """
        row = time_after + FIRST_DATA_ROW

        def get_cell(column: int) -> Cell:
            return sheet.cell(row=row, column=column)

        get_cell(columns.time_after).value = time_after
        get_cell(columns.train1_pax).value = train1_remaining_arrivals + train1_new_pax
        get_cell(columns.train2_pax).value = train2_remaining_arrivals + train2_new_pax
        get_cell(
            columns.arrived_pax_waiting_on_plat
        ).value = arrived_pax_waiting_on_plat
        get_cell(columns.train1_off_rate).value = train1_off_rate
        get_cell(columns.train2_off_rate).value = train2_off_rate
        get_cell(columns.train1_on_rate).value = train1_on_rate
        get_cell(columns.train2_on_rate).value = train2_on_rate
        get_cell(columns.down_rate).value = plat_ingress_rate_1 + plat_ingress_rate_2
        get_cell(columns.departing_pax_on_plat_1).value = train1_boarders_on_plat
        get_cell(columns.departing_pax_on_plat_2).value = train2_boarders_on_plat
        net_pax_flow_rate = (
            plat_ingress_rate_1
            + plat_ingress_rate_2
            + train1_off_rate
            + train2_off_rate
            - plat_egress_rate
            - train1_on_rate
            - train2_on_rate
        )
        get_cell(columns.total_pax_on_platform).value = total_pax_on_platform
        get_cell(columns.inst_crowding).value = inst_crowding
        get_cell(columns.up_rate).value = plat_egress_rate
        get_cell(columns.net_pax_flow_rate).value = net_pax_flow_rate
        get_cell(columns.plat_crowd_los).value = plat_crowd_grade(inst_crowding)

        get_cell(columns.egress_los).value = egress_crowd_grade(
            params.w, plat_egress_rate
        )

    # Time gets exported to column 3, see line 264.
    def make_chart(title, min_col, x_title, y_title):
        chart = openpyxl.chart.ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = params.simulation_time
        chart.legend = None

        max_row = params.simulation_time + FIRST_DATA_ROW - 1
        xvalues = Reference(sheet, min_col=3, min_row=FIRST_DATA_ROW, max_row=max_row)
        values = Reference(
            sheet, min_col=min_col, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        # Y values start one row above X values so that first cell is series name.
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)
        return chart

    def make_chart_with_chopped_y(title, min_col, x_title, y_title):
        chart = openpyxl.chart.ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = params.simulation_time
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 50
        chart.legend = None

        max_row = params.simulation_time + FIRST_DATA_ROW - 1
        xvalues = Reference(sheet, min_col=3, min_row=FIRST_DATA_ROW, max_row=max_row)
        values = Reference(
            sheet, min_col=min_col, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        # Y values start one row above X values so that first cell is series name.
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)
        return chart

    def make_chart_2(title, col1, col2, x_title, y_title):
        chart = openpyxl.chart.ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = params.simulation_time
        chart.legend.position = "b"

        max_row = params.simulation_time + FIRST_DATA_ROW - 1
        xvalues = Reference(sheet, min_col=3, min_row=FIRST_DATA_ROW, max_row=max_row)
        values1 = Reference(
            sheet, min_col=col1, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        values2 = Reference(
            sheet, min_col=col2, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        # Y values start one row above X values so that first cell is series name.
        series1 = Series(values1, xvalues, title_from_data=True)
        chart.series.append(series1)
        series2 = Series(values2, xvalues, title_from_data=True)
        chart.series.append(series2)
        return chart

    sheet.add_chart(
        make_chart_2(
            "Up and Down Rates",
            columns.up_rate,
            columns.down_rate,
            "Time (s)",
            "Rate (pax/s)",
        ),
        "V4",
    )
    sheet.add_chart(
        make_chart_2(
            "Passengers Aboard Trains",
            columns.train1_pax,
            columns.train2_pax,
            "Time (s)",
            "Passengers",
        ),
        "V19",
    )
    sheet.add_chart(
        make_chart_2(
            "Passengers on Platform",
            columns.arrived_pax_waiting_on_plat,
            columns.total_pax_on_platform,
            "Time (s)",
            "Passengers",
        ),
        "V34",
    )
    sheet.add_chart(
        make_chart_with_chopped_y(
            "Space per Passenger",
            columns.inst_crowding,
            "Time (s)",
            "Space per passenger (sqft)",
        ),
        "V49",
    )
    sheet.add_chart(
        make_chart(
            "Net Platform Flow Rate",
            columns.net_pax_flow_rate,
            "Time (s)",
            "Net Flow Rate (pax/s)",
        ),
        "V64",
    )
    print(
        "LOS F egress rate is "
        + str(params.w * 19 / 60)
        + " pax/second. Emergency egress time is roughly "
        + str(
            (params.train1_arriving_pax + params.train2_arriving_pax)
            / (params.w * 19 / 60)
        )
        + " seconds."
    )
    return wb


def main():
    params = Params(
        file_slug="platform3_pr_alt3_new",
        simulation_time=600,
        width=15,
        length=900,
        cf=0.75,
        train1_arriving_pax=1600,
        train2_arriving_pax=1600,
        train1_outbound_demand=400,
        train2_outbound_demand=400,
        train1_awaiting_boarders=200,
        train2_awaiting_boarders=200,
        train1_doors=16,
        train2_doors=16,
        arr_time1=0,
        arr_time2=300,
        queue_length=20,
        w=537 / 12,
        ww=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    wb = calc_workbook(params=params)

    wb.save(
        f"{params.file_slug}_{params.train1_arriving_pax}_{params.train2_arriving_pax}_{params.arr_time2 - params.arr_time1}s.xlsx"
    )
    wb.close()


if __name__ == "__main__":
    main()
