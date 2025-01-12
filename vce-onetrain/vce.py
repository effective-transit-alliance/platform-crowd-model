# This is a recursive platform clearance calculator.
# c is total area of platform minus vce's in sqft.
# k0 is passengers aboard train on arrival
# u is rate of train egress in pax/second; use 1 pax/second/door
# You input desired t1, time elapsed after arrival
# w is total width of upstairs VCEs in feet.
# egress is instantaneous rate of platform egress
# t1 simulation length in seconds

from openpyxl import Workbook

wb = Workbook()

filepath = "platform_F_egress.xlsx"

sheet = wb.active
assert sheet is not None

width = 15
length = 1100
c = 0.75 * width * length
k0 = 2000  # if *2 is here, its a train on both tracks of the platform
u = 48  # if *2 is here, its a train on both tracks of the platform
t1 = 240
w = 550 / 12
egress: float = 0  # do not touch this


def trainloadfn(k: float, u: float, t: float) -> float:
    return max(0, k - u * t)


numonplatform: float = 0

for i in range(1, t1):

    def platloadfn(
        k: float,
        u: float,
        t: float,
        platcnt: float,
    ) -> float:  # number of passengers on platform, determines space per pax
        trnld = trainloadfn(k, u, t)
        if (
            trnld > 0
        ):  # t-30 is a 30 second penalty assume people on train and need to still leave
            return max(0, platcnt + u - egress)
        else:
            return max(0, platcnt - egress)

    def crowdfn(
        k: float, u: float, t: float, c: float, platcnt: float
    ) -> float:  # space per pax, determines egress
        plat = platloadfn(k, u, t, platcnt)
        if plat == 0:
            return c
        else:
            return c / plat

    if i < 60:
        egress = min(
            w * 19 / 60,
            2
            / 60
            * w
            * (111 * crowdfn(k0, u, i - 1, c, numonplatform) - 162)
            / (crowdfn(k0, u, i - 1, c, numonplatform) ** 2),
        )
    else:
        egress = min(
            w * 19 / 60,
            max(
                w * 7 / 60,
                2
                / 60
                * w
                * (111 * crowdfn(k0, u, i - 1, c, numonplatform) - 162)
                / (crowdfn(k0, u, i - 1, c, numonplatform) ** 2),
            ),
        )
        # model from https://onlinepubs.trb.org/Onlinepubs/hrr/1971/355/355-001.pdf
        # we assume flow bottoms out at LOS B/C boundary of 7 pax/min/foot and tops out at LOS F 19 pax/min/foot
    numonplatform = platloadfn(k0, u, i, numonplatform)
    instcrowding = crowdfn(k0, u, i, c, numonplatform)
    instnumontrain = trainloadfn(k0, u, i)
    sheet.cell(row=i + 3, column=1).value = i
    sheet.cell(row=i + 3, column=2).value = instnumontrain
    sheet.cell(row=i + 3, column=3).value = numonplatform
    sheet.cell(row=i + 3, column=4).value = instcrowding
    sheet.cell(row=i + 3, column=6).value = egress

    if instcrowding > 35:
        sheet.cell(row=i + 3, column=5).value = "A"
    elif 25 < instcrowding <= 35:
        sheet.cell(row=i + 3, column=5).value = "B"
    elif 15 < instcrowding <= 25:
        sheet.cell(row=i + 3, column=5).value = "C"
    elif 10 < instcrowding <= 15:
        sheet.cell(row=i + 3, column=5).value = "D"
    elif 5 < instcrowding <= 10:
        sheet.cell(row=i + 3, column=5).value = "E"
    else:
        sheet.cell(row=i + 3, column=5).value = "F"

    if egress <= w * 5 / 60:
        sheet.cell(row=i + 3, column=7).value = "A"
    elif w * 5 / 60 < egress <= w * 7 / 60:
        sheet.cell(row=i + 3, column=7).value = "B"
    elif w * 7 / 60 < egress <= w * 9.5 / 60:
        sheet.cell(row=i + 3, column=7).value = "C"
    elif w * 9.5 / 60 < egress <= w * 13 / 60:
        sheet.cell(row=i + 3, column=7).value = "D"
    elif w * 13 / 60 < egress <= w * 17 / 60:
        sheet.cell(row=i + 3, column=7).value = "E"
    else:
        sheet.cell(row=i + 3, column=7).value = "F"

    print(
        "At time "
        + str(i)
        + ", platform has "
        + str(numonplatform)
        + " passengers, and train has "
        + str(instnumontrain)
        + " passengers. Each passenger has "
        + str(instcrowding)
        + " sf of space. Egress rate is "
        + str(egress)
        + " pax/sec."
    )
print(
    "LOS F egress rate is "
    + str(w * 19 / 60)
    + " pax/min. Emergency egress time is roughly "
    + str(k0 / (w * 19 / 60))
    + " seconds."
)
sheet.cell(row=2, column=1).value = "Platform width"
sheet.cell(row=2, column=2).value = width
sheet.cell(row=2, column=3).value = "Platform length"
sheet.cell(row=2, column=4).value = length
sheet.cell(row=2, column=5).value = "Total VCE width (ft)"
sheet.cell(row=2, column=6).value = w
sheet.cell(row=2, column=7).value = "Arriving Passengers"
sheet.cell(row=2, column=8).value = k0

sheet.cell(row=3, column=1).value = "Time after arrival (s)"
sheet.cell(row=3, column=2).value = "Passengers on train"
sheet.cell(row=3, column=3).value = "Passengers on platform"
sheet.cell(row=3, column=4).value = "Space per passenger (sqft)"
sheet.cell(row=3, column=5).value = "Average Platform LOS"
sheet.cell(row=3, column=6).value = "Egress rate (passengers/sec)"
sheet.cell(row=3, column=7).value = "Stairwell LOS"
wb.save(filepath)
