#This is a recursive platform clearance calculator.
#c is total area of platform minus vce's in sqft.
#k0 is passengers aboard train on arrival
#u is rate of train egress in pax/second; use 1 pax/second/door
#You input desired t1, time elapsed after arrival
#w is total width of upstairs VCEs in feet.
#egress is instantaneous rate of platform egress
#t1 simulation length in seconds
import openpyxl

from openpyxl import Workbook

wb=Workbook()

filepath='platform_egress.xlsx'

sheet=wb.active

width=25
length=1400
c = .9*width*length
k0 = 1600 #if *2 is here, its a train on both tracks of the platform
u = 48 #if *2 is here, its a train on both tracks of the platform
t1 = 200
w = 625/12
egress = 0 #do not touch this

def trainloadfn(k,u,t):
    return max(0,k-u*t)
numonplatform = 0

for i in range (1,t1):
    def platloadfn(k, u, t, platcnt): #number of passengers on platform, determines space per pax
        trnld = trainloadfn(k,u,t)
        if trnld > 0: #t-30 is a 30 second penalty assume people on train and need to still leave
            return max(0, platcnt + u - egress)
        else:
            return max(0,platcnt - egress)
    def crowdfn(k, u, t, c, platcnt): #space per pax, determines egress
        plat = platloadfn(k, u, t, platcnt)
        if plat == 0:
            return c
        else:
            return c / plat
    if i < 60:
        egress = min(w*19/60, 2/60*w*(111 *crowdfn(k0, u, i-1, c, numonplatform)-162)/(crowdfn(k0, u, i-1, c, numonplatform)**2))
    else:
        egress = min(w*19/60, max(w*7/60, 2 / 60 * w * (111 * crowdfn(k0, u, i - 1, c, numonplatform) - 162) / (crowdfn(k0, u, i - 1, c, numonplatform) ** 2)))
        # we assume flow bottoms out at LOS B/C boundary of 7 pax/min/foot and tops out at LOS F 19 pax/min/foot
    numonplatform = platloadfn(k0,u,i, numonplatform)
    instcrowding = crowdfn(k0,u,i, c,numonplatform)
    instnumontrain = trainloadfn(k0,u,i)
    sheet.cell(row = i + 3, column = 1).value=i
    sheet.cell(row=i + 3, column=2).value = instnumontrain
    sheet.cell(row = i+3, column = 3).value = numonplatform
    sheet.cell(row = i + 3, column = 4).value = instcrowding
    sheet.cell(row = i + 3, column = 5).value = egress


    print('At time ' + str(i) + ', platform has ' + str(numonplatform) + ' passengers, and train has ' + str(instnumontrain) + ' passengers. Each passenger has ' + str(instcrowding)  + ' sf of space. Egress rate is ' + str(egress) + ' pax/sec.')
print('LOS F egress rate is ' + str(w*19/60)+ ' pax/min. Emergency egress time is roughly '+ str(k0/(w*19/60))+ ' seconds.')

sheet.cell(row = 2, column = 1).value = 'Time after arrival (s)'
sheet.cell(row = 2, column = 2).value = 'Passengers on train'
sheet.cell(row = 2, column = 3).value = 'Passengers on platform'
sheet.cell(row = 2, column = 4).value = 'Space per passenger (sqft)'
sheet.cell(row = 2, column = 5).value = 'egress rate (passengers/sec)'
wb.save(filepath)
