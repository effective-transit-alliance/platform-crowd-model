#!/usr/bin/env -S uv run

"""
This is a recursive peak-hour platform clearance calculator.
model from https://onlinepubs.trb.org/Onlinepubs/hrr/1971/355/355-001.pdf
"""

from dataclasses import dataclass
from typing import Any
import numpy as np
import openpyxl
from openpyxl.cell import Cell, MergedCell
from openpyxl.chart import Reference
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import ScatterChart
from numpy.typing import NDArray

# basic flow: train egress > platform crowd > VCE egress rate > back to
# platform crowd


# keep high VCE egress rate if queues at stairs are long
def alight_rate_fn(k: float, t: float, t0: float, u: float) -> float:
    """
    :param k: number of people waiting to get off train
    :param t: time pass counter (s)
    :param t0: train arrival time
    :param u: train(x)doors*rate (1 pax/door/s)
    :return: egress rate from train to platform across all doors (pax/s)
    """
    if t > t0:
        return min(k, u)
    else:
        return 0


def plat_clearance_fn(karr: float, a: float, w: float, qmax: float) -> float:
    """
    :param a: usable platform area
    :param w: total width of vertical circulation elements
    :param: karr: number of people waiting to get onto a stairwell
    :param: qmax: number of people that can fit around stair thresholds
    :return: platform egress rate on stairs
    """
    if karr <= qmax:
        return min(
            karr,
            min(17 * w / 60, (111 * a / max(1, karr) - 162) / (a / max(1, karr)) ** 2),
        )
    else:
        return max(
            10 * w / 60,
            # Min of upstairs LOS C/D boundary flow rate
            min(17 * w / 60, (111 * a / max(1, karr) - 162) / (a / max(1, karr)) ** 2),
        )


def plat_ingress_fn(kdep: float, a: float, w: float, r_up: float) -> float:
    """
    :param kdep: number of people waiting to get onto a stairwell
    :param a: usable concourse area
    :param w: total width of vertical circulation elements
    :param: r_up: upstairs flow, passed from plat_egress_fn
    :return: platform ingress rate on stairs
    """
    # 1st question, how much downstairs flow demand exists?
    # 2nd question, how much stair capacity does upstairs flow take?
    # P = (111M - 162)/(M^2) is the upstairs flow eq per ft wide.
    if kdep > 0:
        return min(
            kdep,
            min(
                max(0, 12 * w / 60 - r_up),
                # Max of downstairs LOS C/D boundary flow rate
                max(
                    0,
                    (
                        (111 * (a / max(1, kdep)) - 162) / ((a / max(1, kdep)) ** 2)
                        - r_up
                    ),
                ),
            ),
        )
    else:
        return 0


def boarder_frac_fn(trainA_boarders: float, trainB_boarders: float) -> float:
    if trainA_boarders + trainB_boarders > 0:
        return trainA_boarders / (trainB_boarders + trainA_boarders)
    else:
        return 1


def board_rate_fn(
    r_max: float,
    r_off: float,
    sim_t: float,
    arr_t: float,
    dep_t: float,
    boarders: float,
) -> float:
    """
    :param vmax: maximum train deboard rate,
    pass train1_doors or train2_doors from params, since 1 door/sec
    :param r_off: train alight rate, pass alight_rate_fn
    :param sim_t: time in seconds, pass counter
    :param arr_t: train arrival time
    :param: dep_t: train departure time
    :param: boarders: number of passengers waiting on platform to board,
    pass departing_pax_on_plat
    :return: train ingress rate across all doors (pax/s)
    """
    if arr_t < sim_t < dep_t:
        return min(r_max - r_off, boarders)
    else:
        return 0


def space_per_pax_fn(k: float, a: float) -> float:
    """
    :param k: people on platform (pax)
    :param a: usable platform area (ft^2)
    :return: space per passenger (ft^2/pax)
    """
    if k > 0:
        return a / k
    else:
        return a


def plat_crowd_grade(inst_crowding: float) -> str:
    if inst_crowding > 35:
        return "A"
    elif 25 < inst_crowding <= 35:
        return "B"
    elif 15 < inst_crowding <= 25:
        return "C"
    elif 10 < inst_crowding <= 15:
        return "D"
    elif 5 < inst_crowding <= 10:
        return "E"
    else:
        return "F"


def egress_crowd_grade(w: float, plat_egress_rate: float) -> str:
    if plat_egress_rate <= w * 5 / 60:
        return "A"
    elif w * 5 / 60 < plat_egress_rate <= w * 7 / 60:
        return "B"
    elif w * 7 / 60 < plat_egress_rate <= w * 9.5 / 60:
        return "C"
    elif w * 9.5 / 60 < plat_egress_rate <= w * 13 / 60:
        return "D"
    elif w * 13 / 60 < plat_egress_rate <= w * 17 / 60:
        return "E"
    else:
        return "F"


@dataclass
class Params:
    filename_prefix: str
    """Prefix of filename to save the spreadsheet in."""

    simulation_time: int
    """Time (in seconds) to simulate."""

    platform_width: int
    """Platform width (in feet)."""

    platform_length: int
    """Platform length (in feet)."""

    usable_platform_area_multiplier: float
    """
    A multiplier to estimate the usable platform area (in square feet)
    given obstructive elements on the platform (e.x. stairs, escalators, elevators, columns).
    """

    train1_arriving_pax: int
    """Number of passengers arriving on train 1."""

    train2_arriving_pax: int
    """Number of passengers arriving on train 2."""

    train1_departing_pax: int
    """Number of passengers departing on train 1."""

    train2_departing_pax: int
    """Number of passengers departing on train 2."""

    train1_doors: int
    """Number of doors (single-door equivalents) on train 1."""

    train2_doors: int
    """Number of doors (single-door equivalents) on train 2."""

    train1_arrival_time: int
    """Time (in seconds) when train 1 arrives."""

    train2_arrival_time: int
    """Time (in seconds) when train 2 arrives."""

    queue_length: int
    """Length (in feet) of the queue in front of each stair that pushes max flow."""

    total_vce_width: float
    """Total width (in feet) of all of the VCEs (vertical circulation elements) going upstairs."""

    vce_widths: NDArray[np.floating]
    """Widths (in feet) of each VCE (vertical circulation element)."""

    train1_boarding_pax: int
    """Number of passengers already on the platform at time 0 wanting to board train 1."""

    train2_boarding_pax: int
    """Number of passengers already on the platform at time 0 wanting to board train 2."""


def calc_workbook(params: Params) -> openpyxl.Workbook:
    eff_area = (
        params.platform_width
        * params.platform_length
        * params.usable_platform_area_multiplier
    )

    www = params.vce_widths[0, :]

    print("www = ", www)

    # Initialize counters
    arrived_pax_waiting_on_plat: float = 0
    train1_remaining_arrivals = float(params.train1_arriving_pax)
    train2_remaining_arrivals = float(params.train2_arriving_pax)
    train1_new_pax: float = 0
    train2_new_pax: float = 0
    train1_boarders_upstairs = float(
        params.train1_departing_pax - params.train1_boarding_pax
    )
    train2_boarders_upstairs = float(
        params.train2_departing_pax - params.train2_boarding_pax
    )
    train1_boarders_on_plat = float(params.train1_boarding_pax)
    train2_boarders_on_plat = float(params.train2_boarding_pax)
    total_pax_on_platform = train1_boarders_on_plat + train2_boarders_on_plat
    wb = openpyxl.Workbook()

    assert type(wb.active) is Worksheet
    sheet: Worksheet = wb.active

    rownum = 0

    def make_row(value: Any, description: str) -> None:
        nonlocal rownum
        rownum = rownum + 1
        sheet.cell(column=1, row=rownum).value = description
        sheet.cell(column=2, row=rownum).value = value

    make_row("Value", "Parameter")
    make_row(params.platform_width, "Platform width (ft)")
    make_row(params.platform_length, "Platform length (ft)")
    make_row(params.total_vce_width, "Total VCE width (ft)")
    make_row(params.usable_platform_area_multiplier, "Effective Area Multiplier")
    make_row(eff_area, "Usable Platform Area (sqft)")
    make_row(params.train1_arriving_pax, "Train 1 Arriving Passengers")
    make_row(params.train1_departing_pax, "Train 1 Departing Passengers")
    make_row(params.train1_arrival_time, "Train 1 Arrival Time")
    make_row(params.train2_arriving_pax, "Train 2 Arriving Passengers")
    make_row(params.train2_departing_pax, "Train 2 Departing Passengers")
    make_row(params.train2_arrival_time, "Train 2 Arrival Time")
    make_row(params.simulation_time, "Simulation Length (s)")
    make_row(params.total_vce_width * 19 / 60, "LOS F Egress Rate (pax/s)")
    make_row(
        (params.train1_arriving_pax + params.train2_arriving_pax)
        / (params.total_vce_width * 19 / 60),
        "Emergency Egress Time (s)",
    )

    del rownum

    @dataclass
    class Columns:
        time_after: int
        train1_pax: int
        train2_pax: int
        train1_off_rate: int
        train2_off_rate: int
        train1_on_rate: int
        train2_on_rate: int
        down_rate: int
        up_rate: int
        departing_pax_on_plat_1: int
        departing_pax_on_plat_2: int
        arrived_pax_waiting_on_plat: int
        total_pax_on_platform: int
        inst_crowding: int
        net_pax_flow_rate: int
        plat_crowd_los: int
        egress_los: int

    # The input parameters go in a table taking up columns 0 (A) and 1 (B).
    colnum = 2

    def make_column_num(description: str) -> int:
        nonlocal colnum
        colnum = colnum + 1
        sheet.cell(row=1, column=colnum).value = description
        return colnum

    columns = Columns(
        time_after=make_column_num("Time after arrival (s),"),
        train1_pax=make_column_num("Passengers on Train 1"),
        train2_pax=make_column_num("Passengers on Train 2"),
        train1_off_rate=make_column_num("Train 1 Alight Rate (pax/s),"),
        train2_off_rate=make_column_num("Train 2 Alight Rate (pax/s),"),
        train1_on_rate=make_column_num("Train 1 Board Rate (pax/s),"),
        train2_on_rate=make_column_num("Train 2 Board Rate (pax/s),"),
        down_rate=make_column_num("Downstairs Rate (pax/s),"),
        up_rate=make_column_num("Upstairs Rate (pax/s),"),
        departing_pax_on_plat_1=make_column_num(
            "Train 1 Departing Passengers on Platform"
        ),
        departing_pax_on_plat_2=make_column_num(
            "Train 2 Departing Passengers on Platform"
        ),
        arrived_pax_waiting_on_plat=make_column_num("Arrived Passengers on Platform"),
        total_pax_on_platform=make_column_num("Total Passengers on Platform"),
        inst_crowding=make_column_num("Platform Space per Passanger (sqft),"),
        net_pax_flow_rate=make_column_num("Net Platform Flow Rate"),
        plat_crowd_los=make_column_num("Platform Crowding LOS"),
        egress_los=make_column_num("Egress LOS"),
    )

    del colnum

    FIRST_DATA_ROW = 2

    print("Elapsed_Time", "Train_1_Pax", "Train_2_Pax")

    for time_after in range(0, params.simulation_time):
        train1_off_rate = alight_rate_fn(
            train1_remaining_arrivals,
            time_after,
            params.train1_arrival_time,
            params.train1_doors,
        )
        train1_remaining_arrivals -= train1_off_rate
        if train1_remaining_arrivals < 0:
            train1_remaining_arrivals = 0
        train2_off_rate = alight_rate_fn(
            train2_remaining_arrivals,
            time_after,
            params.train2_arrival_time,
            params.train2_doors,
        )
        train2_remaining_arrivals -= train2_off_rate
        if train2_remaining_arrivals < 0:
            train2_remaining_arrivals = 0
        total_pax_on_platform += train1_off_rate + train2_off_rate
        arrived_pax_waiting_on_plat += train1_off_rate + train2_off_rate
        plat_egress_rate = plat_clearance_fn(
            arrived_pax_waiting_on_plat,
            eff_area,
            params.total_vce_width,
            params.total_vce_width * params.queue_length / 5,
        )
        arrived_pax_waiting_on_plat -= plat_egress_rate
        if arrived_pax_waiting_on_plat < 0:
            arrived_pax_waiting_on_plat = 0
        total_pax_on_platform -= plat_egress_rate
        plat_ingress_rate_1 = plat_ingress_fn(
            train1_boarders_upstairs,
            5000,
            params.total_vce_width
            * boarder_frac_fn(train1_boarders_upstairs, train2_boarders_upstairs),
            plat_egress_rate,
        )

        plat_ingress_rate_2 = plat_ingress_fn(
            train2_boarders_upstairs,
            5000,
            params.total_vce_width
            * boarder_frac_fn(train2_boarders_upstairs, train1_boarders_upstairs),
            plat_egress_rate,
        )
        train1_boarders_on_plat += plat_ingress_rate_1
        train2_boarders_on_plat += plat_ingress_rate_2
        total_pax_on_platform += plat_ingress_rate_1
        total_pax_on_platform += plat_ingress_rate_2
        train1_on_rate = board_rate_fn(
            params.train1_doors,
            train1_off_rate,
            time_after,
            params.train1_arrival_time,
            params.simulation_time,
            train1_boarders_on_plat,
        )
        train2_on_rate = board_rate_fn(
            params.train2_doors,
            train2_off_rate,
            time_after,
            params.train2_arrival_time,
            params.simulation_time,
            train2_boarders_on_plat,
        )

        train1_boarders_on_plat -= train1_on_rate

        train2_boarders_on_plat -= train2_on_rate

        total_pax_on_platform -= train1_on_rate

        total_pax_on_platform -= train2_on_rate

        train1_boarders_upstairs -= plat_ingress_rate_1

        train2_boarders_upstairs -= plat_ingress_rate_2

        train1_new_pax += train1_on_rate

        train2_new_pax += train2_on_rate

        inst_crowding = space_per_pax_fn(total_pax_on_platform, eff_area)
        if total_pax_on_platform < 0:
            total_pax_on_platform = 0
        if train1_boarders_on_plat < 0:
            train1_boarders_on_plat = 0
        if train2_boarders_on_plat < 0:
            train2_boarders_on_plat = 0
        if arrived_pax_waiting_on_plat < 0:
            arrived_pax_waiting_on_plat = 0
        print(
            time_after,
            train1_remaining_arrivals + train1_new_pax,
            train2_remaining_arrivals + train2_new_pax,
            arrived_pax_waiting_on_plat,
            plat_egress_rate,
        )
        """
        print(
            "At time " + str(time_after) + " s,",
            str(train1_remaining_arrivals) + " wait to alight train 1;",
            str(train2_remaining_arrivals) + " wait to alight train 2;",
            str(train1_on_rate - train1_off_rate) + " pax/s train 1 net rate;",
            str(train2_on_rate - train2_off_rate) + " pax/s train 2 net rate;",
        )
        print(
            str(int(arrived_pax_waiting_on_plat))
            + " deboarded pax on platform;",
            str(int(train1_boarders_on_plat + int(train2_boarders_on_plat)))
            + " boarding pax on platform;",
            str(int(total_pax_on_platform)) + " total pax on platform;",
            str(int(inst_crowding)) + " sqft per pax;",
        )
        print(
            str(plat_egress_rate) + " pax/s up;",
            str((plat_ingress_rate_1 + plat_ingress_rate_2)) + "pax/s down;",
            str((train1_boarders_upstairs + train2_boarders_upstairs))
            + " pax are upstairs"
        )
        """
        row = time_after + FIRST_DATA_ROW

        def get_cell(column: int) -> Cell | MergedCell:
            return sheet.cell(row=row, column=column)

        get_cell(columns.time_after).value = time_after
        get_cell(columns.train1_pax).value = train1_remaining_arrivals + train1_new_pax
        get_cell(columns.train2_pax).value = train2_remaining_arrivals + train2_new_pax
        get_cell(
            columns.arrived_pax_waiting_on_plat
        ).value = arrived_pax_waiting_on_plat
        get_cell(columns.train1_off_rate).value = train1_off_rate
        get_cell(columns.train2_off_rate).value = train2_off_rate
        get_cell(columns.train1_on_rate).value = train1_on_rate
        get_cell(columns.train2_on_rate).value = train2_on_rate
        get_cell(columns.down_rate).value = plat_ingress_rate_1 + plat_ingress_rate_2
        get_cell(columns.departing_pax_on_plat_1).value = train1_boarders_on_plat
        get_cell(columns.departing_pax_on_plat_2).value = train2_boarders_on_plat
        net_pax_flow_rate = (
            plat_ingress_rate_1
            + plat_ingress_rate_2
            + train1_off_rate
            + train2_off_rate
            - plat_egress_rate
            - train1_on_rate
            - train2_on_rate
        )
        get_cell(columns.total_pax_on_platform).value = total_pax_on_platform
        get_cell(columns.inst_crowding).value = inst_crowding
        get_cell(columns.up_rate).value = plat_egress_rate
        get_cell(columns.net_pax_flow_rate).value = net_pax_flow_rate
        get_cell(columns.plat_crowd_los).value = plat_crowd_grade(inst_crowding)

        get_cell(columns.egress_los).value = egress_crowd_grade(
            params.total_vce_width, plat_egress_rate
        )

    # Time gets exported to column 3, see line 264.
    def make_chart(
        title: str, min_col: int, x_title: str, y_title: str
    ) -> ScatterChart:
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = params.simulation_time
        chart.legend = None

        max_row = params.simulation_time + FIRST_DATA_ROW - 1
        xvalues = Reference(sheet, min_col=3, min_row=FIRST_DATA_ROW, max_row=max_row)
        values = Reference(
            sheet, min_col=min_col, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        # Y values start one row above X values so that first cell is series name.
        series = SeriesFactory(values, xvalues, title_from_data=True)
        chart.series.append(series)
        return chart

    def make_chart_with_chopped_y(
        title: str, min_col: int, x_title: str, y_title: str
    ) -> ScatterChart:
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = params.simulation_time
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 50
        chart.legend = None

        max_row = params.simulation_time + FIRST_DATA_ROW - 1
        xvalues = Reference(sheet, min_col=3, min_row=FIRST_DATA_ROW, max_row=max_row)
        values = Reference(
            sheet, min_col=min_col, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        # Y values start one row above X values so that first cell is series name.
        series = SeriesFactory(values, xvalues, title_from_data=True)
        chart.series.append(series)
        return chart

    def make_chart_2(
        title: str, col1: int, col2: int, x_title: str, y_title: str
    ) -> ScatterChart:
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = params.simulation_time
        assert chart.legend is not None
        chart.legend.position = "b"

        max_row = params.simulation_time + FIRST_DATA_ROW - 1
        xvalues = Reference(sheet, min_col=3, min_row=FIRST_DATA_ROW, max_row=max_row)
        values1 = Reference(
            sheet, min_col=col1, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        values2 = Reference(
            sheet, min_col=col2, min_row=FIRST_DATA_ROW - 1, max_row=max_row
        )
        # Y values start one row above X values so that first cell is series name.
        series1 = SeriesFactory(values1, xvalues, title_from_data=True)
        chart.series.append(series1)
        series2 = SeriesFactory(values2, xvalues, title_from_data=True)
        chart.series.append(series2)
        return chart

    sheet.add_chart(
        make_chart_2(
            "Up and Down Rates",
            columns.up_rate,
            columns.down_rate,
            "Time (s)",
            "Rate (pax/s)",
        ),
        "V4",
    )
    sheet.add_chart(
        make_chart_2(
            "Passengers Aboard Trains",
            columns.train1_pax,
            columns.train2_pax,
            "Time (s)",
            "Passengers",
        ),
        "V19",
    )
    sheet.add_chart(
        make_chart_2(
            "Passengers on Platform",
            columns.arrived_pax_waiting_on_plat,
            columns.total_pax_on_platform,
            "Time (s)",
            "Passengers",
        ),
        "V34",
    )
    sheet.add_chart(
        make_chart_with_chopped_y(
            "Space per Passenger",
            columns.inst_crowding,
            "Time (s)",
            "Space per passenger (sqft)",
        ),
        "V49",
    )
    sheet.add_chart(
        make_chart(
            "Net Platform Flow Rate",
            columns.net_pax_flow_rate,
            "Time (s)",
            "Net Flow Rate (pax/s)",
        ),
        "V64",
    )
    print(
        "LOS F egress rate is "
        + str(params.total_vce_width * 19 / 60)
        + " pax/second. Emergency egress time is roughly "
        + str(
            (params.train1_arriving_pax + params.train2_arriving_pax)
            / (params.total_vce_width * 19 / 60)
        )
        + " seconds."
    )
    return wb


def run_model(params: Params) -> None:
    wb = calc_workbook(params=params)

    wb.save(
        f"{params.filename_prefix}_{params.train1_arriving_pax}_{params.train2_arriving_pax}_{params.train2_arrival_time - params.train1_arrival_time}s.xlsx"
    )
    wb.close()


def main() -> None:
    # params are labeled  with p<platform number><time in seconds> recon indicates that a platform was modelled accounting for penn reconstruction plans
    params_p3120 = Params(
        filename_prefix="platform3",
        simulation_time=600,
        platform_width=18,
        platform_length=900,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=120,
        queue_length=20,
        total_vce_width=42.5,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    params_p3300 = Params(
        filename_prefix="platform3",
        simulation_time=600,
        platform_width=18,
        platform_length=900,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=300,
        queue_length=20,
        total_vce_width=42.5,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    params_p3recon120 = Params(
        filename_prefix="platform3_recon",
        simulation_time=600,
        platform_width=18,
        platform_length=900,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=120,
        queue_length=20,
        total_vce_width=44.75,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    params_p3recon300 = Params(
        filename_prefix="platform3_recon",
        simulation_time=600,
        platform_width=18,
        platform_length=900,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=300,
        queue_length=20,
        total_vce_width=44.75,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    params_p60 = Params(
        filename_prefix="platform6",
        simulation_time=600,
        platform_width=15,
        platform_length=1100,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=0,
        queue_length=20,
        total_vce_width=48.168,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    params_p10120 = Params(
        filename_prefix="platform10",
        simulation_time=600,
        platform_width=42,
        platform_length=1100,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=120,
        queue_length=20,
        total_vce_width=70.58,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    params_p11120 = Params(
        filename_prefix="platform11",
        simulation_time=600,
        platform_width=18,
        platform_length=1100,
        usable_platform_area_multiplier=0.75,
        train1_arriving_pax=1620,
        train2_arriving_pax=1620,
        train1_departing_pax=400,
        train2_departing_pax=400,
        train1_boarding_pax=200,
        train2_boarding_pax=200,
        train1_doors=40,
        train2_doors=40,
        train1_arrival_time=0,
        train2_arrival_time=120,
        queue_length=20,
        total_vce_width=43.58,
        vce_widths=(
            1
            / 12
            * np.transpose(
                np.array(
                    [
                        [60, 1],
                        [60, 1],
                        [40, 1],
                        [54, 1],
                        [40, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                        [54, 1],
                    ]
                )
            )
        ),
    )
    run_model(params_p3120)
    run_model(params_p3300)
    run_model(params_p3recon120)
    run_model(params_p3recon300)
    run_model(params_p60)
    run_model(params_p10120)
    run_model(params_p11120)


if __name__ == "__main__":
    main()
