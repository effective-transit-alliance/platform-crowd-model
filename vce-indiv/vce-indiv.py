#This is a recursive peak-hour platform clearance calculator.
#simtime simulation length in seconds
#c is total area of platform minus vce's in sqft.
#k1 is passengers aboard train 1 on arrival; k2 is train 2
#u1 is rate of train 1 egress, u2 train 2 egress in pax/second; use 1 pax/second/door
#t1 and t2, are times at which trains 1 and 2 arrive
#w is total width of upstairs VCEs in feet.
#egress is instantaneous rate of platform egress
# model from https://onlinepubs.trb.org/Onlinepubs/hrr/1971/355/355-001.pdf
import openpyxl
import numpy as np
filepath='platform_F_twoway.xlsx'

wb = openpyxl.load_workbook(filepath)

sheet=wb.active

simtime = 600  #how many seconds we want to simulate
width=15
length=1100
cf = .75  #area correction factor
eff_area = width*length*cf
k1 = 1600  #train 1 load
k2 = 1600  #train 2 load
u1 = 48  #train 1 max egress rate. Use 1 pax/second/single door, 2 pax/second/double door.
u2 = 48  #train 2 max egress rate. Use 1 pax/second/single door, 2 pax/second/double door.
t1 = 0  #time of first train arrival
t2 = 150  #time of second train arrival
w = 550/12
ww = 1/12*np.array([[60, 60, 36, 36, 40, 54, 40, 64, 54,54, 54, 54, 54],[0.6,0.6,0.6,0.6,1.0,1.0,1.0,1.0,1.0,1.4,1.4,1.4,1.4]])
www = ww[0,:]
print(www)
egress = 0  #initialize platform egress at 0
ingress = 0

#basic flow: train egress > platform crowd > VCE egress rate > back to platform crowd
# if t2 >= simtime, only consider one train.

def uncortrainloadfn(k,u,t, arrtime): # of people on train 1 without ingress
    if t >= arrtime:
        return max(0, k - u*(t - arrtime))
    else:
        return k

numonplatform = 0  #initialize number of pax on platform at zero, or desired preexisting load
train1pax = k1
train2pax = k2

for i in range(0, simtime):

    if i >= t1:
        if train1pax > 0:
            alightrate1 = u1
        else:
            alightrate1 = 0
    else:
        alightrate1 = 0

    if i >= t2:
        if train2pax > 0:
            alightrate2 = u2
        else:
            alightrate2 = 0
    else:
        alightrate2 = 0


    def uncorplatloadfn(k, u, t, platcnt):  #gives current platform pax count, determines space per pax. platcnt is prior  pax count.
        return max(0, platcnt + alightrate1 + alightrate2 - egress)


    def crowdfn(k, u, t, eff_area, platcnt):  # space per pax, function of pax arriving on trains, egress rates, time, area, platcnt; determines egress
        plat = uncorplatloadfn(k1 + k2, alightrate1 + alightrate2, t, platcnt)
        if plat == 0:
            return eff_area
        else:
            return eff_area / plat


    if i < 60:
        egress = min(w * 19 / 60, 2 / 60 * w * (111 * crowdfn(k1 + k2, alightrate1 + alightrate2, i - 1, eff_area, numonplatform) - 162) / (
                    crowdfn(k1 + k2, alightrate1 + alightrate2, i - 1, eff_area, numonplatform) ** 2))
    else:
        egress = min(w * 19 / 60, max(w * 7 / 60, 2 / 60 * w * (111 * crowdfn(k1 + k2, alightrate1 + alightrate2, i - 1, eff_area, numonplatform) - 162) / (
                    crowdfn(k1 + k2, alightrate1 + alightrate2, i - 1, eff_area, numonplatform) ** 2)))
    numonplatform = uncorplatloadfn(k1 + k2, alightrate1 + alightrate2, i, numonplatform)

    def ingressfn(exit):
        if w*7/60 <= exit <= w * 15/60:
            return w * 0.2/60
        elif exit < w * 7/60:
            return w * 0.5/60
        else:
            return w*0.05/60
    instingress = ingressfn(egress)
    numonplatform = numonplatform + instingress

    instcrowding = crowdfn(k1 + k2, alightrate1 + alightrate2, i, eff_area, numonplatform)
    train1pax = uncortrainloadfn(k1, u1, i, t1)
    train2pax = uncortrainloadfn(k2, u2, i, t2)

    print('time ' + str(i), 'train 1 load ' + str(train1pax), 'train 2 load ' + str(train2pax), alightrate1, alightrate2, numonplatform, instcrowding, egress)
    sheet.cell(row = i+3, column = 1).value = i
    sheet.cell(row=i + 3, column=2).value = train1pax
    sheet.cell(row=i + 3, column=3).value = train2pax
    sheet.cell(row=i+3, column=4).value = alightrate1
    sheet.cell(row = i+3, column = 5).value = alightrate2
    sheet.cell(row = i+3, column = 6).value = numonplatform
    sheet.cell(row = i+3, column = 7).value = instcrowding
    sheet.cell(row=i + 3, column=8).value = egress
    egr = egress / width * www
    print(egr, np.sum(egr))
    if instcrowding > 35:
        sheet.cell(row=i + 3, column=9).value = 'A'
    elif 25 < instcrowding <= 35:
        sheet.cell(row=i + 3, column=9).value = 'B'
    elif 15 < instcrowding <= 25:
        sheet.cell(row=i + 3, column=9).value = 'C'
    elif 10 < instcrowding <= 15:
        sheet.cell(row=i + 3, column=9).value = 'D'
    elif 5 < instcrowding <= 10:
        sheet.cell(row=i + 3, column=9).value = 'E'
    else:
        sheet.cell(row=i + 3, column=9).value = 'F'

    if egress <= w * 5 / 60:
        sheet.cell(row=i + 3, column=10).value = 'A'
    elif w * 5 / 60 < egress <= w * 7 / 60:
        sheet.cell(row=i + 3, column=10).value = 'B'
    elif w * 7 / 60 < egress <= w * 9.5 / 60:
        sheet.cell(row=i + 3, column=10).value = 'C'
    elif w * 9.5 / 60 < egress <= w * 13 / 60:
        sheet.cell(row=i + 3, column=10).value = 'D'
    elif w * 13 / 60 < egress <= w * 17 / 60:
        sheet.cell(row=i + 3, column=10).value = 'E'
    else:
        sheet.cell(row=i + 3, column=10).value = 'F'

sheet.cell(row = 2, column = 1).value = 'Platform width (ft)'
sheet.cell(row = 2, column = 2).value = width
sheet.cell(row = 2, column = 3).value = 'Platform length (ft)'
sheet.cell(row = 2, column = 4).value = length
sheet.cell(row = 2, column = 5).value = 'Total VCE width (ft)'
sheet.cell(row = 2, column = 6).value = w
sheet.cell(row = 2, column = 7).value = 'Effective Area Multiplier'
sheet.cell(row = 2, column = 8).value = cf
sheet.cell(row = 2, column = 9).value = 'Usable Platform Area (sqft)'
sheet.cell(row = 2, column = 10).value = eff_area
sheet.cell(row = 2, column = 11).value = 'Train 1 Arriving Passengers'
sheet.cell(row = 2, column = 12).value = k1
sheet.cell(row = 2, column = 13).value = 'Train 1 Arrival Time'
sheet.cell(row = 2, column = 14).value = t1
sheet.cell(row = 2, column = 15).value = 'Train 2 Arriving Passengers'
sheet.cell(row = 2, column = 16).value = k2
sheet.cell(row = 2, column = 17).value = 'Train 2 Arrival Time'
sheet.cell(row = 2, column = 18).value = t2
sheet.cell(row = 2, column = 19).value = 'Simulation Length (s)'
sheet.cell(row = 2, column = 20).value = simtime

sheet.cell(row = 3, column = 1).value = 'Time after arrival (s)'
sheet.cell(row = 3, column = 2).value = 'Passengers on Train 1'
sheet.cell(row = 3, column = 3).value = 'Passengers on Train 2'
sheet.cell(row = 3, column = 4).value = 'Train 1 Alight Rate (pax/s)'
sheet.cell(row = 3, column = 5).value = 'Train 2 Alight Rate (pax/s)'
sheet.cell(row = 3, column = 6).value = 'Passengers on Platform'
sheet.cell(row = 3, column = 7).value = 'Average Space on Platform per Passenger (sqft)'
sheet.cell(row = 3, column = 8).value = 'Platform Egress Rate (pax/s)'
sheet.cell(row = 3, column = 9).value = 'Platform Crowding LOS'
sheet.cell(row = 3, column = 10).value = 'Egress LOS'

print('LOS F egress rate is ' + str(w*19/60)+ ' pax/min. Emergency egress time is roughly '+ str((k1 + k2)/(w*19/60))+ ' seconds.')
wb.save(filepath)